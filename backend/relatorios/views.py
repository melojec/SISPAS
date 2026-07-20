import html
import io
import re
from datetime import date

from django.conf import settings
from django.db.models.expressions import RawSQL
from django.http import HttpResponse
from django.template.loader import render_to_string
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.exceptions import NotFound
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from core.models import Meta, Area
from monitoramento.models import RegistroQuadrimestral, Ciclo
from usuarios.permissions import IsUsuarioAtivo


def _nat(table):
    engine = settings.DATABASES['default']['ENGINE']
    if 'sqlite' in engine:
        return ['codigo']
    if 'postgresql' in engine:
        return [
            RawSQL(f'CAST(SPLIT_PART("{table}"."codigo", \'.\', 1) AS INTEGER)', []),
            RawSQL(f'CAST(SPLIT_PART("{table}"."codigo", \'.\', 2) AS INTEGER)', []),
            RawSQL(f'CAST(SPLIT_PART("{table}"."codigo", \'.\', 3) AS INTEGER)', []),
        ]
    return [
        RawSQL(f"CAST(SUBSTRING_INDEX(`{table}`.`codigo`, '.', 1) AS UNSIGNED)", []),
        RawSQL(f"CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(`{table}`.`codigo`, '.', 2), '.', -1) AS UNSIGNED)", []),
        RawSQL(f"CAST(SUBSTRING_INDEX(`{table}`.`codigo`, '.', -1) AS UNSIGNED)", []),
    ]


class MetaPDFView(APIView):
    permission_classes = [IsUsuarioAtivo]

    def get(self, request, meta_id):
        try:
            meta = Meta.objects.select_related(
                'area', 'objetivo', 'objetivo__diretriz'
            ).prefetch_related('atividades').get(pk=meta_id)
        except Meta.DoesNotExist:
            raise NotFound('Meta não encontrada.')

        ciclo_id = request.query_params.get('ciclo')
        ciclo = Ciclo.objects.filter(pk=ciclo_id).first() if ciclo_id else None

        registros_qs = RegistroQuadrimestral.objects.filter(
            meta=meta
        ).select_related('ciclo').order_by('ciclo__ano', 'ciclo__quadrimestre')

        # Monta lista fixa de 3 quadrimestres com valor realizado (0 se sem registro)
        reg_por_q = {r.ciclo.quadrimestre: r for r in registros_qs if r.ciclo}
        labels_q = {1: '1º Quadrimestre', 2: '2º Quadrimestre', 3: '3º Quadrimestre'}
        valores_realizados = [
            {
                'label': labels_q[q],
                'valor': reg_por_q[q].realizado if q in reg_por_q else 0,
                'registro': reg_por_q.get(q),
            }
            for q in [1, 2, 3]
        ]

        # Registro do ciclo atual para campos qualitativos
        registro_atual = reg_por_q.get(ciclo.quadrimestre) if ciclo else None

        import os, base64
        _logo_candidates = [
            (os.path.join(os.path.dirname(__file__), 'logo_pdf.png'), 'image/png'),
            (os.path.join(os.path.dirname(__file__), 'logo.png'), 'image/png'),
            (os.path.join(os.path.dirname(__file__), 'logo.svg'), 'image/svg+xml'),
        ]
        logo_path = ''
        for _candidate, _mime in _logo_candidates:
            if os.path.exists(_candidate):
                with open(_candidate, 'rb') as f:
                    logo_path = f'data:{_mime};base64,{base64.b64encode(f.read()).decode()}'
                break

        html_string = render_to_string('relatorios/meta_pdf.html', {
            'meta': meta,
            'ciclo': ciclo,
            'valores_realizados': valores_realizados,
            'registro_atual': registro_atual,
            'logo_path': logo_path,
            'data_geracao': date.today().strftime('%d/%m/%Y'),
        })
        from xhtml2pdf import pisa
        buffer = io.BytesIO()
        pisa.CreatePDF(html_string, dest=buffer)
        buffer.seek(0)
        nome = f'meta_{meta.codigo.replace(".", "_")}.pdf'
        return HttpResponse(buffer.read(), content_type='application/pdf',
                            headers={'Content-Disposition': f'attachment; filename="{nome}"'})


class TodasMetasPDFView(APIView):
    permission_classes = [IsUsuarioAtivo]

    def get(self, request):
        import traceback
        try:
            return self._get(request)
        except Exception as e:
            import json
            body = json.dumps({'erro': f'{type(e).__name__}: {e}', 'detalhe': traceback.format_exc()})
            return HttpResponse(body, content_type='application/json', status=500)

    def _get(self, request):
        import os
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        pt = 1  # 1 point = 1 ReportLab unit
        from reportlab.platypus import (
            Paragraph, Table, TableStyle, Spacer, PageBreak, SimpleDocTemplate,
            KeepTogether, Image,
        )
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY

        ciclo_id = request.query_params.get('ciclo')
        area_id  = request.query_params.get('area')
        ciclo       = Ciclo.objects.filter(pk=ciclo_id).first() if ciclo_id else None
        area_filtro = Area.objects.filter(pk=area_id).first()  if area_id  else None

        metas_qs = Meta.objects.select_related(
            'area', 'objetivo', 'objetivo__diretriz'
        ).prefetch_related('atividades').order_by(*_nat('meta'))
        if area_id:
            metas_qs = metas_qs.filter(area_id=area_id)

        registros_all = RegistroQuadrimestral.objects.filter(
            meta__in=metas_qs
        ).select_related('ciclo').order_by('ciclo__ano', 'ciclo__quadrimestre')
        regs_por_meta = {}
        regs_atual = {}
        for r in registros_all:
            regs_por_meta.setdefault(r.meta_id, {})[r.ciclo.quadrimestre] = r
            if ciclo:
                if r.ciclo_id == ciclo.pk:
                    regs_atual[r.meta_id] = r
            else:
                regs_atual[r.meta_id] = r

        # ── Logo ─────────────────────────────────────────────────────────
        _logo_dir = os.path.dirname(__file__)
        logo_file = None
        for _name in ['logo_pdf.png', 'logo.png', 'logo.svg']:
            _p = os.path.join(_logo_dir, _name)
            if os.path.exists(_p):
                logo_file = _p
                break

        def _make_logo(size=52):
            if not logo_file:
                return Spacer(size, size)
            if logo_file.endswith('.svg'):
                try:
                    from svglib.svglib import svg2rlg
                    from reportlab.graphics import renderPDF
                    drawing = svg2rlg(logo_file)
                    scale = (size * pt) / max(drawing.width, drawing.height)
                    drawing.width  = drawing.width  * scale
                    drawing.height = drawing.height * scale
                    drawing.transform = (scale, 0, 0, scale, 0, 0)
                    return drawing
                except Exception:
                    return Spacer(size, size)
            return Image(logo_file, width=size * pt, height=size * pt)

        # ── Cores ─────────────────────────────────────────────────────────
        AZUL       = colors.HexColor('#172554')
        AZUL_H     = colors.HexColor('#A2B1D9')   # header de tabela
        AZUL_LABEL = colors.HexColor('#dbeafe')   # fundo labels análise
        AZUL_TINT  = colors.HexColor('#f8faff')   # fundo painel direito / linhas pares
        AZUL_VR    = colors.HexColor('#eff6ff')   # fundo célula "Valores Realizados"
        BORDA      = colors.HexColor('#172554')
        BORDA_INTR = colors.HexColor('#c5cfe8')
        CINZA_TEXT = colors.HexColor('#374151')
        LABEL_TEXT = colors.HexColor('#6b7280')

        # ── Estilos de texto ──────────────────────────────────────────────
        def st(name, **kw):
            base = dict(fontName='Helvetica', fontSize=9, leading=12,
                        textColor=colors.black, spaceAfter=0, spaceBefore=0)
            base.update(kw)
            return ParagraphStyle(name, **base)

        sNormal  = st('n', fontSize=9, leading=13, textColor=CINZA_TEXT)
        sSmallGr = st('sgr', fontSize=8, leading=11, textColor=CINZA_TEXT)
        sLabel   = st('lbl', fontSize=7.5, leading=10, textColor=LABEL_TEXT,
                      fontName='Helvetica-Bold', wordWrap='CJK')
        sWhiteBd = st('wb', fontSize=10.5, fontName='Helvetica-Bold',
                      textColor=colors.white, leading=15, alignment=TA_CENTER)
        sWhiteSm = st('wsm', fontSize=8, textColor=colors.white, leading=11)
        sTitle   = st('title', fontSize=13, fontName='Helvetica-Bold',
                      textColor=AZUL, leading=16)
        sSub     = st('sub', fontSize=8, textColor=LABEL_TEXT, leading=10)
        sRight   = st('right', fontSize=8, textColor=CINZA_TEXT,
                      leading=11, alignment=TA_RIGHT)
        sAreaLbl = st('albl', fontSize=7.5, textColor=LABEL_TEXT,
                      fontName='Helvetica-Bold', alignment=TA_CENTER,
                      wordWrap='CJK')
        sAreaNome= st('anome', fontSize=10.5, fontName='Helvetica-Bold',
                      textColor=colors.black, alignment=TA_CENTER, leading=14)
        sPlanLbl = st('plbl', fontSize=8.5, textColor=LABEL_TEXT, alignment=TA_CENTER)
        sPlanVal = st('pval', fontSize=22, fontName='Helvetica-Bold',
                      textColor=AZUL, alignment=TA_CENTER, leading=26)
        sVRLbl   = st('vrlbl', fontSize=7.5, fontName='Helvetica-Bold',
                      textColor=AZUL, leading=10)
        sQLbl    = st('ql', fontSize=8.5, textColor=CINZA_TEXT, alignment=TA_CENTER)
        sQVal    = st('qv', fontSize=22, fontName='Helvetica-Bold',
                      textColor=AZUL, alignment=TA_CENTER, leading=26)
        sAtvTh   = st('ath', fontSize=8.5, fontName='Helvetica-Bold',
                      textColor=colors.HexColor('#172457'), alignment=TA_CENTER)
        sAtvTd   = st('atd', fontSize=9.5, textColor=colors.black, leading=13)
        sAqLbl   = st('aql', fontSize=9.5, fontName='Helvetica-Bold',
                      textColor=colors.HexColor('#1e3a5f'))
        sAqText  = st('aqt', fontSize=8.5, textColor=colors.black,
                      leading=13, alignment=TA_JUSTIFY)
        sAqEmpty = st('aqe', fontSize=8.5, textColor=LABEL_TEXT, fontName='Helvetica-Oblique')
        sBcLeft  = st('bcl', fontSize=9.5, textColor=CINZA_TEXT, leading=14)

        def _strip_html(text):
            if not text:
                return ''
            return re.sub(r'<[^>]+>', '', html.unescape(str(text))).strip()

        def _fmt(val, unidade=''):
            if val is None:
                return '—'
            try:
                v = float(val)
                if 'porcentagem' in (unidade or '').lower():
                    return f'{v:,.2f}%'.replace(',', 'X').replace('.', ',').replace('X', '.')
                return f'{int(round(v))}'
            except Exception:
                return str(val)

        # ── Dimensões ─────────────────────────────────────────────────────
        W       = A4[0] - 32*mm
        W_AREA  = 46*mm
        W_PLAN  = 62*mm
        W_VRL   = 22*mm
        LOGO_SZ = 14*mm   # 52pt ≈ 14mm
        data_geracao = date.today().strftime('%d/%m/%Y')

        def _cabecalho(ciclo_obj):
            ciclo_txt = f'Ciclo: {ciclo_obj}<br/>' if ciclo_obj else ''
            logo = _make_logo(14)
            data = [[
                logo,
                [Paragraph('Ficha de Meta', sTitle),
                 Paragraph('Sistema de Monitoramento da Programação Anual de Saúde — SES-MA', sSub)],
                Paragraph(f'{ciclo_txt}Gerado em: {data_geracao}', sRight),
            ]]
            t = Table(data, colWidths=[LOGO_SZ + 2*mm, W - LOGO_SZ - 2*mm - 38*mm, 38*mm])
            t.setStyle(TableStyle([
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                ('TOPPADDING',    (0,0), (-1,-1), 3),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('LEFTPADDING',   (0,0), (-1,-1), 0),
                ('RIGHTPADDING',  (0,0), (-1,-1), 0),
                ('LINEBELOW',     (0,0), (-1,-1), 2, AZUL),
            ]))
            return t

        # ── Monta story ───────────────────────────────────────────────────
        story = []

        for idx, meta in enumerate(metas_qs):
            reg_por_q      = regs_por_meta.get(meta.pk, {})
            registro_atual = regs_atual.get(meta.pk)
            un = meta.unidade or ''
            d   = meta.objetivo.diretriz
            obj = meta.objetivo

            if idx > 0:
                story.append(PageBreak())

            # ── Cabeçalho ─────────────────────────────────────────────────
            story.append(_cabecalho(ciclo))
            story.append(Spacer(1, 3*mm))

            # ── Breadcrumb + Área ─────────────────────────────────────────
            bc_left = [
                Paragraph(f'<b>{_strip_html(d.codigo)}</b> - {_strip_html(d.descricao)}', sBcLeft),
                Paragraph(f'<b>{_strip_html(obj.codigo)}</b> - {_strip_html(obj.descricao)}', sBcLeft),
            ]
            bc_right = [
                Paragraph('Área Responsável', sAreaLbl),
                Paragraph(
                    f'{meta.area.sigla} — {meta.area.nome}' if meta.area else '—',
                    sAreaNome
                ),
            ]
            bc = Table([[bc_left, bc_right]], colWidths=[W - W_AREA, W_AREA])
            bc.setStyle(TableStyle([
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                ('TOPPADDING',    (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4*mm),
                ('LEFTPADDING',   (0,0), (0,-1), 0),
                ('LEFTPADDING',   (1,0), (1,-1), 4),
                ('RIGHTPADDING',  (0,0), (0,-1), 8),
                ('RIGHTPADDING',  (1,0), (1,-1), 0),
                ('BOX',           (1,0), (1,-1), 1, BORDA),
                ('TOPPADDING',    (1,0), (1,-1), 6),
                ('BOTTOMPADDING', (1,0), (1,-1), 6),
            ]))
            story.append(bc)

            # ── Banner meta ───────────────────────────────────────────────
            banner_data = [[Paragraph(
                f'{_strip_html(meta.codigo)} - {_strip_html(meta.descricao)}', sWhiteBd
            )]]
            story.append(Table(banner_data, colWidths=[W],
                style=TableStyle([
                    ('BACKGROUND',    (0,0), (-1,-1), AZUL),
                    ('TOPPADDING',    (0,0), (-1,-1), 6),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                    ('LEFTPADDING',   (0,0), (-1,-1), 8),
                    ('RIGHTPADDING',  (0,0), (-1,-1), 8),
                ])))

            # ── Indicador + Valores Planejados ────────────────────────────
            ind_txt = _strip_html(meta.indicador) or ''
            un_txt  = _strip_html(meta.unidade)   or ''
            ind_content = []
            if ind_txt:
                ind_content.append(Paragraph(f'<b>Indicador:</b> {ind_txt}', sNormal))
            if un_txt:
                ind_content.append(Paragraph(f'<b>Unidade:</b> {un_txt}', sNormal))
            if not ind_content:
                ind_content.append(Paragraph('<i>Não informado</i>', st('ni', textColor=LABEL_TEXT, fontName='Helvetica-Oblique')))

            ppa_lbl  = 'PES (4 anos)'
            pas_lbl  = f'PAS (Ano {ciclo.ano})' if ciclo else 'PAS'
            ppa_val  = _fmt(meta.previsto_ppa, un)
            exer_val = _fmt(meta.previsto_exercicio, un)

            iv_right = [
                Paragraph('Valores Planejados', sLabel),
                Table(
                    [[Paragraph(ppa_lbl, sPlanLbl), Paragraph(pas_lbl, sPlanLbl)],
                     [Paragraph(ppa_val, sPlanVal), Paragraph(exer_val, sPlanVal)]],
                    colWidths=[(W_PLAN - 4*mm) / 2] * 2,
                    style=TableStyle([
                        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
                        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                        ('TOPPADDING',    (0,0), (-1,-1), 2),
                        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
                        ('LINEAFTER',     (0,0), (0,-1), 0.5, colors.HexColor('#e5e7eb')),
                    ])
                ),
            ]
            iv = Table([[ind_content, iv_right]], colWidths=[W - W_PLAN, W_PLAN])
            iv.setStyle(TableStyle([
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                ('TOPPADDING',    (0,0), (-1,-1), 7),
                ('BOTTOMPADDING', (0,0), (-1,-1), 7),
                ('LEFTPADDING',   (0,0), (0,-1), 8),
                ('LEFTPADDING',   (1,0), (1,-1), 6),
                ('RIGHTPADDING',  (0,0), (-1,-1), 6),
                ('BOX',           (0,0), (-1,-1), 1, BORDA),
                ('LINEAFTER',     (0,0), (0,-1), 1, BORDA),
                ('BACKGROUND',    (1,0), (1,-1), AZUL_TINT),
                ('LINEBEFORE',    (0,0), (-1,-1), 0, colors.white),
            ]))
            story.append(iv)

            # ── Valores Realizados ────────────────────────────────────────
            q_labels = ['1º Quadrimestre', '2º Quadrimestre', '3º Quadrimestre']
            q_vals   = [_fmt(reg_por_q[q].realizado, un) if q in reg_por_q else '—'
                        for q in [1, 2, 3]]
            vr_cells = [Paragraph('Valores\nRealizados', sVRLbl)]
            for lbl, val in zip(q_labels, q_vals):
                vr_cells.append([Paragraph(lbl, sQLbl), Paragraph(val, sQVal)])

            vr = Table([vr_cells],
                colWidths=[W_VRL] + [(W - W_VRL) / 3] * 3)
            vr.setStyle(TableStyle([
                ('BACKGROUND',    (0,0), (0,-1), AZUL_VR),
                ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN',         (1,0), (-1,-1), 'CENTER'),
                ('TOPPADDING',    (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('LEFTPADDING',   (0,0), (0,-1), 8),
                ('LEFTPADDING',   (1,0), (-1,-1), 4),
                ('BOX',           (0,0), (-1,-1), 1, BORDA),
                ('LINEAFTER',     (0,0), (0,-1), 1, BORDA),
                ('LINEAFTER',     (1,0), (1,-1), 0.5, BORDA_INTR),
                ('LINEAFTER',     (2,0), (2,-1), 0.5, BORDA_INTR),
            ]))
            story.append(vr)
            story.append(Spacer(1, 4*mm))

            # ── Atividades ────────────────────────────────────────────────
            atividades = list(meta.atividades.all())
            if atividades:
                atv_banner = Table([[Paragraph('Atividades', sWhiteBd)]], colWidths=[W],
                    style=TableStyle([
                        ('BACKGROUND',    (0,0), (-1,-1), AZUL),
                        ('TOPPADDING',    (0,0), (-1,-1), 6),
                        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                        ('LEFTPADDING',   (0,0), (-1,-1), 8),
                    ]))
                story.append(atv_banner)

                w_ind  = 36*mm
                w_un   = 18*mm
                w_meta = 13*mm
                w_desc = W - w_ind - w_un - w_meta
                atv_rows = [[
                    Paragraph('Descrição',              sAtvTh),
                    Paragraph('Indicador da Atividade', sAtvTh),
                    Paragraph('Unidade',                sAtvTh),
                    Paragraph('Meta',                   sAtvTh),
                ]]
                for i, a in enumerate(atividades):
                    bg = AZUL_TINT if i % 2 == 1 else colors.white
                    atv_rows.append([
                        Paragraph(_strip_html(a.descricao), sAtvTd),
                        Paragraph(_strip_html(a.indicador) or '—', st(f'ati{i}', fontSize=9.5, alignment=TA_CENTER)),
                        Paragraph(_strip_html(a.unidade)   or '—', st(f'atu{i}', fontSize=9.5, alignment=TA_CENTER)),
                        Paragraph(_fmt(a.valor_previsto, un), st(f'atm{i}', fontSize=9.5, alignment=TA_CENTER)),
                    ])

                atv_tbl = Table(atv_rows, colWidths=[w_desc, w_ind, w_un, w_meta])
                row_styles = [
                    ('BACKGROUND',    (0,0), (-1,0), AZUL_H),
                    ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
                    ('ALIGN',         (1,0), (-1,-1), 'CENTER'),
                    ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
                    ('TOPPADDING',    (0,0), (-1,-1), 4),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                    ('LEFTPADDING',   (0,0), (-1,-1), 6),
                    ('RIGHTPADDING',  (0,0), (-1,-1), 6),
                    ('BOX',           (0,0), (-1,-1), 1, BORDA_INTR),
                    ('INNERGRID',     (0,0), (-1,-1), 0.5, BORDA_INTR),
                ]
                for i in range(1, len(atv_rows)):
                    if i % 2 == 0:
                        row_styles.append(('BACKGROUND', (0,i), (-1,i), AZUL_TINT))
                atv_tbl.setStyle(TableStyle(row_styles))
                story.append(atv_tbl)

            # ── Pág. 2 — Análise Qualitativa ─────────────────────────────
            story.append(PageBreak())
            story.append(_cabecalho(ciclo))
            story.append(Spacer(1, 3*mm))

            aq_banner = Table(
                [[Paragraph(f'Análise Qualitativa', sWhiteBd)]],
                colWidths=[W],
                style=TableStyle([
                    ('BACKGROUND',    (0,0), (-1,-1), AZUL),
                    ('TOPPADDING',    (0,0), (-1,-1), 7),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 7),
                    ('LEFTPADDING',   (0,0), (-1,-1), 8),
                ])
            )
            story.append(aq_banner)
            story.append(Spacer(1, 4*mm))

            def _aq_field(label, valor):
                lbl_row = Table([[Paragraph(label, sAqLbl)]], colWidths=[W],
                    style=TableStyle([
                        ('BACKGROUND',    (0,0), (-1,-1), AZUL_LABEL),
                        ('TOPPADDING',    (0,0), (-1,-1), 4),
                        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                        ('LEFTPADDING',   (0,0), (-1,-1), 8),
                    ]))
                txt = _strip_html(valor) if valor else ''
                txt_para = (Paragraph(txt, sAqText) if txt
                            else Paragraph('Não informado', sAqEmpty))
                return [lbl_row, txt_para, Spacer(1, 4*mm)]

            story += _aq_field('Problemas Encontrados no Quadrimestre',
                               registro_atual.problema if registro_atual else None)
            story += _aq_field('Ações Realizadas para o Enfrentamento dos Problemas',
                               registro_atual.acao if registro_atual else None)
            story += _aq_field('Análises e Considerações – Este texto irá diretamente para o DigiSUS',
                               registro_atual.analise if registro_atual else None)

        # ── Gera PDF ──────────────────────────────────────────────────────
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
            leftMargin=16*mm, rightMargin=16*mm,
            topMargin=14*mm, bottomMargin=16*mm)
        doc.build(story)
        buffer.seek(0)
        return HttpResponse(buffer.read(), content_type='application/pdf',
            headers={'Content-Disposition': 'attachment; filename="fichas_metas.pdf"'})


class RelatorioPDFView(APIView):
    permission_classes = [IsUsuarioAtivo]

    def get(self, request):
        import traceback, json
        try:
            return self._get(request)
        except Exception as e:
            body = json.dumps({'erro': f'{type(e).__name__}: {e}', 'detalhe': traceback.format_exc()})
            return HttpResponse(body, content_type='application/json', status=500)

    def _get(self, request):
        ciclo_id = request.query_params.get('ciclo')
        area_id = request.query_params.get('area')

        qs = RegistroQuadrimestral.objects.select_related(
            'meta', 'meta__area', 'meta__objetivo', 'meta__objetivo__diretriz', 'ciclo'
        )
        if ciclo_id:
            qs = qs.filter(ciclo_id=ciclo_id)
        if area_id:
            qs = qs.filter(meta__area_id=area_id)

        ciclo = Ciclo.objects.filter(pk=ciclo_id).first() if ciclo_id else None
        html_string = render_to_string('relatorios/relatorio_pas.html', {
            'registros': qs,
            'ciclo': ciclo,
        })
        from xhtml2pdf import pisa
        buffer = io.BytesIO()
        pisa.CreatePDF(html_string, dest=buffer)
        buffer.seek(0)
        return HttpResponse(buffer.read(), content_type='application/pdf',
                            headers={'Content-Disposition': 'attachment; filename="relatorio_pas.pdf"'})


class RelatorioXLSXView(APIView):
    permission_classes = [IsUsuarioAtivo]

    def get(self, request):
        import traceback, json
        try:
            return self._get(request)
        except Exception as e:
            body = json.dumps({'erro': f'{type(e).__name__}: {e}', 'detalhe': traceback.format_exc()})
            return HttpResponse(body, content_type='application/json', status=500)

    def _get(self, request):
        ciclo_id = request.query_params.get('ciclo')
        area_id = request.query_params.get('area')

        qs = RegistroQuadrimestral.objects.select_related(
            'meta', 'meta__area', 'meta__objetivo', 'meta__objetivo__diretriz', 'ciclo'
        )
        if ciclo_id:
            qs = qs.filter(ciclo_id=ciclo_id)
        if area_id:
            qs = qs.filter(meta__area_id=area_id)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Registros PAS'

        cabecalho = [
            'Diretriz', 'Objetivo', 'Meta (Código)', 'Meta (Descrição)',
            'Área', 'Ciclo', 'Previsto (Exercício)', 'Realizado',
            'Problema', 'Ação', 'Análise', 'Valid. Coord.', 'Valid. ASPLAN',
        ]
        header_fill = PatternFill('solid', fgColor='1a3a5c')
        header_font = Font(bold=True, color='FFFFFF')

        for col, titulo in enumerate(cabecalho, 1):
            cell = ws.cell(row=1, column=col, value=titulo)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row, reg in enumerate(qs, 2):
            ws.cell(row=row, column=1, value=reg.meta.objetivo.diretriz.codigo)
            ws.cell(row=row, column=2, value=reg.meta.objetivo.codigo)
            ws.cell(row=row, column=3, value=reg.meta.codigo)
            ws.cell(row=row, column=4, value=reg.meta.descricao)
            ws.cell(row=row, column=5, value=reg.meta.area.nome)
            ws.cell(row=row, column=6, value=str(reg.ciclo))
            ws.cell(row=row, column=7, value=float(reg.meta.previsto_exercicio or 0))
            ws.cell(row=row, column=8, value=float(reg.realizado or 0))
            ws.cell(row=row, column=9, value=reg.problema)
            ws.cell(row=row, column=10, value=reg.acao)
            ws.cell(row=row, column=11, value=reg.analise)
            ws.cell(row=row, column=12, value='Sim' if reg.validado_coord else 'Não')
            ws.cell(row=row, column=13, value='Sim' if reg.validado_asplan else 'Não')

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="relatorio_pas.xlsx"'},
        )
