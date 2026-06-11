"""
Script de importação da Base PAS.xlsx para o banco sispas_dev.
Execute com: python importar_pas.py
"""
import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from datetime import date
from core.models import Area, Diretriz, Objetivo, Meta, Atividade
from monitoramento.models import Ciclo

ARQUIVO = '../Base PAS.xlsx'

print('Abrindo planilha...')
wb = openpyxl.load_workbook(ARQUIVO, read_only=True, data_only=True)

# ── 1. Área placeholder ───────────────────────────────────────────────────────
area, _ = Area.objects.get_or_create(
    sigla='GERAL',
    defaults={'nome': 'A Definir', 'ativa': True}
)
print(f'Área: {area}')

# ── 2. Diretrizes ─────────────────────────────────────────────────────────────
ws = wb['Diretriz']
diretrizes = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    num, descricao = row[0], row[1]
    if not num:
        continue
    d, criado = Diretriz.objects.update_or_create(
        codigo=str(int(num)),
        defaults={'descricao': descricao or '', 'situacao': True, 'ano': 2026}
    )
    diretrizes[int(num)] = d
    print(f'  {"+" if criado else "~"} Diretriz {d.codigo}')

# ── 3. Objetivos ──────────────────────────────────────────────────────────────
ws = wb['Objetivo']
objetivos = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    num_dir, _, codigo, descricao = row[0], row[1], row[2], row[3]
    if not codigo:
        continue
    diretriz = diretrizes.get(int(num_dir))
    if not diretriz:
        continue
    o, criado = Objetivo.objects.update_or_create(
        codigo=str(codigo),
        defaults={'diretriz': diretriz, 'descricao': descricao or ''}
    )
    objetivos[str(codigo)] = o
    print(f'  {"+" if criado else "~"} Objetivo {o.codigo}')

# ── 4. Metas ──────────────────────────────────────────────────────────────────
ws = wb['Meta']
metas = {}
sem_objetivo = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    num_dir, cod_obj, cod_meta, descricao, indicador, unidade, previsto_ppa, previsto_ano = row
    if not cod_meta:
        continue
    objetivo = objetivos.get(str(cod_obj))
    if not objetivo:
        sem_objetivo += 1
        continue
    m, criado = Meta.objects.update_or_create(
        codigo=str(cod_meta),
        defaults={
            'objetivo': objetivo,
            'area': area,
            'descricao': descricao or '',
            'indicador': indicador or '',
            'unidade': unidade or '',
            'previsto_ppa': float(previsto_ppa or 0),
            'previsto_exercicio': float(previsto_ano or 0),
        }
    )
    metas[str(cod_meta)] = m

total_metas = Meta.objects.count()
print(f'  Metas importadas: {total_metas} (ignoradas sem objetivo: {sem_objetivo})')

# ── 5. Atividades ─────────────────────────────────────────────────────────────
ws = wb['Atividade']
atividades_criadas = 0
atividades_sem_meta = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    _, _, _, _, cod_meta, _, cod_ativ, descricao, _, _, valor_meta = row
    if not cod_ativ or not descricao:
        continue
    meta = metas.get(str(cod_meta))
    if not meta:
        atividades_sem_meta += 1
        continue
    try:
        valor = float(valor_meta) if valor_meta not in (None, 'NULL', '') else 0
    except (ValueError, TypeError):
        valor = 0
    _, criado = Atividade.objects.update_or_create(
        meta=meta,
        descricao=str(descricao),
        defaults={'valor_previsto': valor}
    )
    if criado:
        atividades_criadas += 1

total_ativ = Atividade.objects.count()
print(f'  Atividades importadas: {total_ativ} (ignoradas sem meta: {atividades_sem_meta})')

# ── 6. Ciclo atual ────────────────────────────────────────────────────────────
ciclo, criado = Ciclo.objects.get_or_create(
    ano=2026,
    quadrimestre=2,
    defaults={
        'dt_abertura': date(2026, 5, 1),
        'dt_encerramento': date(2026, 8, 31),
        'situacao': Ciclo.ABERTO,
    }
)
print(f'\n{"+" if criado else "~"} Ciclo: {ciclo}')

print('\n✓ Importação concluída.')
print(f'  Diretrizes : {Diretriz.objects.count()}')
print(f'  Objetivos  : {Objetivo.objects.count()}')
print(f'  Metas      : {Meta.objects.count()}')
print(f'  Atividades : {Atividade.objects.count()}')
print(f'  Ciclos     : {Ciclo.objects.count()}')
