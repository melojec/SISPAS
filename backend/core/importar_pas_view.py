import io
import traceback
import openpyxl
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework import status
from usuarios.permissions import IsAdministrador
from .models import Area, Diretriz, Objetivo, Meta, Atividade


def _to_float(value, unidade=''):
    try:
        v = float(value or 0)
    except (ValueError, TypeError):
        v = 0.0
    if v != 0 and 'porcentagem' in str(unidade or '').lower():
        v = round(v * 100, 10)
    return v


class ImportarPASView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsAdministrador]

    def post(self, request):
        arquivo = request.FILES.get('arquivo')
        ano = request.data.get('ano')

        if not arquivo:
            return Response({'erro': 'Arquivo não enviado.'}, status=status.HTTP_400_BAD_REQUEST)
        if not ano:
            return Response({'erro': 'Ano não informado.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            ano = int(ano)
        except (ValueError, TypeError):
            return Response({'erro': 'Ano inválido.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(io.BytesIO(arquivo.read()), read_only=True, data_only=True)
        except Exception as e:
            return Response({'erro': f'Erro ao abrir planilha: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        abas_necessarias = {'Diretriz', 'Objetivo', 'Meta', 'Atividade'}
        abas_faltando = abas_necessarias - set(wb.sheetnames)
        if abas_faltando:
            return Response({'erro': f'Abas não encontradas: {", ".join(sorted(abas_faltando))}. Abas disponíveis: {", ".join(sorted(wb.sheetnames))}'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            resultado = {}

            area, _ = Area.objects.get_or_create(
                sigla='GERAL',
                defaults={'nome': 'A Definir', 'ativa': True}
            )

            # Diretrizes
            diretrizes = {}
            criadas_d = atualizadas_d = 0
            for row in wb['Diretriz'].iter_rows(min_row=2, values_only=True):
                num, descricao = row[0], row[1]
                if not num:
                    continue
                d, criado = Diretriz.objects.update_or_create(
                    codigo=str(int(num)),
                    ano=ano,
                    defaults={'descricao': descricao or '', 'situacao': True}
                )
                diretrizes[int(num)] = d
                if criado:
                    criadas_d += 1
                else:
                    atualizadas_d += 1
            resultado['diretrizes'] = {'criadas': criadas_d, 'atualizadas': atualizadas_d}

            # Objetivos
            objetivos = {}
            criados_o = atualizados_o = 0
            for row in wb['Objetivo'].iter_rows(min_row=2, values_only=True):
                num_dir, _, codigo, descricao = row[0], row[1], row[2], row[3]
                if not codigo:
                    continue
                diretriz = diretrizes.get(int(num_dir)) if num_dir else None
                if not diretriz:
                    continue
                o, criado = Objetivo.objects.update_or_create(
                    codigo=str(codigo),
                    defaults={'diretriz': diretriz, 'descricao': descricao or ''}
                )
                objetivos[str(codigo)] = o
                if criado:
                    criados_o += 1
                else:
                    atualizados_o += 1
            resultado['objetivos'] = {'criados': criados_o, 'atualizados': atualizados_o}

            # Metas
            metas = {}
            criadas_m = atualizadas_m = ignoradas_m = 0
            for row in wb['Meta'].iter_rows(min_row=2, values_only=True):
                num_dir, cod_obj, cod_meta, descricao, indicador, unidade, previsto_ppa, previsto_ano, *_ = row
                if not cod_meta:
                    continue
                objetivo = objetivos.get(str(cod_obj)) if cod_obj else None
                if not objetivo:
                    ignoradas_m += 1
                    continue
                m, criado = Meta.objects.update_or_create(
                    codigo=str(cod_meta),
                    defaults={
                        'objetivo': objetivo,
                        'area': area,
                        'descricao': descricao or '',
                        'indicador': indicador or '',
                        'unidade': unidade or '',
                        'previsto_ppa': _to_float(previsto_ppa, unidade),
                        'previsto_exercicio': _to_float(previsto_ano, unidade),
                    }
                )
                metas[str(cod_meta)] = m
                if criado:
                    criadas_m += 1
                else:
                    atualizadas_m += 1
            resultado['metas'] = {'criadas': criadas_m, 'atualizadas': atualizadas_m, 'ignoradas': ignoradas_m}

            # Atividades
            criadas_a = atualizadas_a = ignoradas_a = 0
            for row in wb['Atividade'].iter_rows(min_row=2, values_only=True):
                _, _, _, _, cod_meta, _, cod_ativ, descricao, indicador_at, unidade_at, valor_meta, *_ = row
                if not cod_ativ or not descricao:
                    continue
                meta = metas.get(str(cod_meta)) if cod_meta else None
                if not meta:
                    ignoradas_a += 1
                    continue
                valor = _to_float(valor_meta if valor_meta not in (None, 'NULL', '') else 0, unidade_at)
                _, criado = Atividade.objects.update_or_create(
                    meta=meta,
                    codigo=str(cod_ativ),
                    defaults={
                        'descricao': str(descricao),
                        'valor_previsto': valor,
                        'indicador': str(indicador_at) if indicador_at else '',
                        'unidade': str(unidade_at) if unidade_at else '',
                    }
                )
                if criado:
                    criadas_a += 1
                else:
                    atualizadas_a += 1
            resultado['atividades'] = {'criadas': criadas_a, 'atualizadas': atualizadas_a, 'ignoradas': ignoradas_a}

            return Response({'sucesso': True, 'ano': ano, 'resultado': resultado})

        except Exception as e:
            tb = traceback.format_exc()
            return Response({'erro': f'{type(e).__name__}: {e}', 'detalhe': tb}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
